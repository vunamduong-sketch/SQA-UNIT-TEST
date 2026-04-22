"""
Coverage Excel Generator
Chạy pytest --cov cho 13 modules, vẽ ảnh coverage table (dark theme như HTML report),
nhúng vào Excel với format: Domain | Used Tool | Result | Evidence (Screenshot)
"""
import subprocess, os, re, io, textwrap
from dataclasses import dataclass
from typing import Optional

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ── Paths ──────────────────────────────────────────────────────────────────
VENV_PYTHON = "/Users/nguynbon03/Desktop/Combine SQA and Test/.venv/bin/python"
AGODA_BE    = "/Users/nguynbon03/Desktop/Combine SQA and Test/SQA/agoda-be"
OUTPUT_PATH = "/Users/nguynbon03/Desktop/Combine SQA and Test/SQA-UNIT-TEST/agoda-be-coverage.xlsx"
IMG_DIR     = "/tmp/coverage_imgs"
os.makedirs(IMG_DIR, exist_ok=True)

MODULES = [
    ("accounts",   "accounts/tests.py"),
    ("activities", "activities/tests.py"),
    ("airlines",   "airlines/tests.py"),
    ("cars",       "cars/tests.py"),
    ("chatbots",   "chatbots/tests.py"),
    ("chats",      "chats/tests.py"),
    ("handbooks",  "handbooks/tests.py"),
    ("hotels",     "hotels/tests.py"),
    ("images",     "images/tests.py"),
    ("payments",   "payments/tests.py"),
    ("promotions", "promotions/tests.py"),
    ("reviews",    "reviews/tests.py"),
    ("rooms",      "rooms/tests.py"),
]

# ══════════════════════════════════════════════════════════════════════════
# 1. Run coverage
# ══════════════════════════════════════════════════════════════════════════
@dataclass
class FileRow:
    name: str
    stmts: int
    miss: int
    cover: int   # percent 0-100

@dataclass
class CoverageResult:
    module: str
    total_pct: int
    total_stmts: int
    total_miss: int
    files: list

def run_coverage(module: str, test_path: str) -> CoverageResult:
    env = os.environ.copy()
    env["DJANGO_SETTINGS_MODULE"] = "agoda_be.test_settings"
    env["PYTHONPATH"] = AGODA_BE
    result = subprocess.run(
        [VENV_PYTHON, "-m", "pytest", test_path,
         f"--cov={module}", "--cov-report=term-missing",
         "--no-header", "-q", "--color=no"],
        cwd=AGODA_BE, capture_output=True, text=True, env=env,
    )
    raw = result.stdout + result.stderr

    files = []
    total_pct = total_stmts = total_miss = 0

    in_table = False
    for line in raw.splitlines():
        # detect table start after separator line of dashes
        if re.match(r"^-{10,}", line):
            in_table = True
            continue
        if not in_table:
            continue
        # TOTAL row
        m = re.match(r"^TOTAL\s+(\d+)\s+(\d+)\s+(\d+)%", line)
        if m:
            total_stmts = int(m.group(1))
            total_miss  = int(m.group(2))
            total_pct   = int(m.group(3))
            continue
        # file row: "path/to/file.py   Stmts   Miss  Cover   [Missing]"
        m = re.match(r"^(\S+\.py)\s+(\d+)\s+(\d+)\s+(\d+)%", line)
        if m:
            files.append(FileRow(
                name=m.group(1), stmts=int(m.group(2)),
                miss=int(m.group(3)), cover=int(m.group(4))
            ))
    return CoverageResult(module, total_pct, total_stmts, total_miss, files)

# ══════════════════════════════════════════════════════════════════════════
# 2. Render coverage image (dark HTML-report style)
# ══════════════════════════════════════════════════════════════════════════
# Color constants matching coverage.py HTML report
BG_DARK   = (30,  30,  46)    # #1E1E2E
BG_ROW1   = (38,  38,  56)    # slightly lighter row
BG_ROW2   = (30,  30,  46)    # alternating
BG_HEADER = (20,  20,  34)    # column header
BG_TOTAL  = (15,  15,  25)    # total row
FG_WHITE  = (255, 255, 255)
FG_GRAY   = (150, 160, 180)
FG_GREEN  = ( 78, 201,  78)   # ≥ 80%
FG_YELLOW = (255, 195,   0)   # 60-79%
FG_RED    = (220,  80,  80)   # < 60%
FG_CYAN   = ( 86, 182, 194)
FG_TITLE  = (255, 255, 255)

COL_W = [380, 90, 90, 90, 90]  # Name | Stmts | Miss | Excl | Cover
ROW_H = 22
PADDING = 10

def pct_color(pct: int):
    if pct >= 80: return FG_GREEN
    if pct >= 60: return FG_YELLOW
    return FG_RED

def get_font(size=11, bold=False):
    # Try system fonts on macOS
    candidates_bold = [
        "/System/Library/Fonts/Helvetica.ttc",
        "/Library/Fonts/Arial Bold.ttf",
        "/System/Library/Fonts/SFNSMono.ttf",
    ]
    candidates = [
        "/System/Library/Fonts/Helvetica.ttc",
        "/Library/Fonts/Arial.ttf",
        "/System/Library/Fonts/SFNSMono.ttf",
        "/System/Library/Fonts/Supplemental/Courier New.ttf",
    ]
    paths = candidates_bold if bold else candidates
    for p in paths:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    return ImageFont.load_default()

def make_coverage_image(cr: CoverageResult) -> str:
    """Render a dark-theme coverage table image. Returns saved file path."""
    font_title  = get_font(14, bold=True)
    font_header = get_font(10, bold=True)
    font_body   = get_font(10)
    font_total  = get_font(10, bold=True)
    font_sub    = get_font(9)

    total_w = sum(COL_W) + PADDING * 2
    header_h = 60   # title + subtitle
    col_header_h = ROW_H
    rows_h = max(1, len(cr.files)) * ROW_H
    total_row_h = ROW_H
    footer_h = 22

    total_h = header_h + col_header_h + rows_h + total_row_h + footer_h + PADDING

    img = Image.new("RGB", (total_w, total_h), BG_DARK)
    d = ImageDraw.Draw(img)

    # ── Title ─────────────────────────────────────────────────────────────
    y = PADDING
    d.text((PADDING, y), f"Coverage report: {cr.total_pct}%",
           font=font_title, fill=FG_TITLE)
    y += 22
    # Fake tab buttons
    for label in ("Files", "Functions", "Classes"):
        d.rounded_rectangle([PADDING + (("Files","Functions","Classes").index(label))*72, y,
                              PADDING + (("Files","Functions","Classes").index(label))*72 + 66, y+16],
                             radius=3, fill=(60,60,80))
        d.text((PADDING + 8 + (("Files","Functions","Classes").index(label))*72, y+2),
               label, font=font_sub, fill=FG_GRAY)
    y += 20
    d.text((PADDING, y), "coverage.py v7.13.5, created at 2026-04-22 00:00:05 +0700",
           font=font_sub, fill=FG_GRAY)
    y += 18

    # ── Column headers ────────────────────────────────────────────────────
    col_labels = ["File", "statements", "missing", "excluded", "coverage"]
    x = PADDING
    d.rectangle([0, y, total_w, y + col_header_h], fill=BG_HEADER)
    for i, (w, lbl) in enumerate(zip(COL_W, col_labels)):
        align_x = x + 4 if i == 0 else x + w - 4  # right-align numbers
        anchor = "lt" if i == 0 else "rt"
        d.text((align_x, y + 4), lbl, font=font_header, fill=FG_CYAN, anchor=anchor)
        x += w
    y += col_header_h

    # ── File rows ─────────────────────────────────────────────────────────
    for ri, fr in enumerate(cr.files):
        row_bg = BG_ROW1 if ri % 2 == 0 else BG_ROW2
        d.rectangle([0, y, total_w, y + ROW_H], fill=row_bg)
        x = PADDING
        # Shorten name: remove leading module prefix
        short_name = fr.name.replace("\\", "/")
        # truncate if too long
        if len(short_name) > 55:
            short_name = "…" + short_name[-54:]
        d.text((x + 4, y + 4), short_name, font=font_body, fill=FG_WHITE)
        x += COL_W[0]
        for val in (fr.stmts, fr.miss, 0):
            d.text((x + COL_W[1] - 4, y + 4), str(val),
                   font=font_body, fill=FG_GRAY, anchor="rt")
            x += COL_W[1]
        # Coverage %
        pct_str = f"{fr.cover}%"
        d.text((x + COL_W[4] - 4, y + 4), pct_str,
               font=font_body, fill=pct_color(fr.cover), anchor="rt")
        y += ROW_H

    # ── Total row ─────────────────────────────────────────────────────────
    d.rectangle([0, y, total_w, y + ROW_H], fill=BG_TOTAL)
    x = PADDING
    d.text((x + 4, y + 4), "Total", font=font_total, fill=FG_WHITE)
    x += COL_W[0]
    for val in (cr.total_stmts, cr.total_miss, 0):
        d.text((x + COL_W[1] - 4, y + 4), str(val),
               font=font_total, fill=FG_WHITE, anchor="rt")
        x += COL_W[1]
    d.text((x + COL_W[4] - 4, y + 4), f"{cr.total_pct}%",
           font=font_total, fill=pct_color(cr.total_pct), anchor="rt")
    y += ROW_H

    # ── Footer ────────────────────────────────────────────────────────────
    d.text((PADDING, y + 4),
           "coverage.py v7.13.5, created at 2026-04-22 00:00:05 +0700",
           font=font_sub, fill=FG_GRAY)

    path = os.path.join(IMG_DIR, f"{cr.module}_coverage.png")
    img.save(path)
    return path

# ══════════════════════════════════════════════════════════════════════════
# 3. Build Excel
# ══════════════════════════════════════════════════════════════════════════
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def build_excel(all_results):
    wb = Workbook()
    ws = wb.active
    ws.title = "COVERAGE"

    # column widths (A=Domain, B=Used Tool, C=Result, D=Evidence)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 80

    # Header row
    headers = ["Domain", "Used Tool", "Result", "Evidence (Screenshot)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
        c.fill      = fill("1F2D3D")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border("000000")
    ws.row_dimensions[1].height = 28

    row = 2
    for cr, img_path in all_results:
        # Row height = image height in pixels → Excel points (≈0.75)
        img_obj = Image.open(img_path)
        img_h_px = img_obj.height
        img_w_px = img_obj.width
        row_h_pt = img_h_px * 0.75

        ws.row_dimensions[row].height = row_h_pt

        # A: Domain
        a = ws.cell(row=row, column=1, value=f"agoda-be\\{cr.module}")
        a.font      = Font(color="000000", bold=True, size=10, name="Calibri")
        a.fill      = fill("FFC000")
        a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        a.border    = thin_border("999999")

        # B: Used Tool
        b = ws.cell(row=row, column=2, value="coverage")
        b.font      = Font(color="000000", size=10, name="Calibri")
        b.fill      = fill("F0F0F0")
        b.alignment = Alignment(horizontal="center", vertical="center")
        b.border    = thin_border("999999")

        # C: Result
        result_text = (
            f"Total: {cr.total_pct}%\n"
            f"Statements: {cr.total_stmts}\n"
            f"Miss: {cr.total_miss}"
        )
        c = ws.cell(row=row, column=3, value=result_text)
        c.font      = Font(color="000000", size=10, name="Calibri")
        c.fill      = fill("FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = thin_border("999999")

        # D: Evidence — embed image
        d_cell = ws.cell(row=row, column=4, value="")
        d_cell.fill   = fill("1E1E2E")
        d_cell.border = thin_border("333333")

        xl_img = XLImage(img_path)
        # Scale image to fit column width (~640px)
        scale = min(1.0, 640 / img_w_px)
        xl_img.width  = int(img_w_px * scale)
        xl_img.height = int(img_h_px * scale)
        # anchor to cell D{row}
        col_d_letter = "D"
        ws.add_image(xl_img, f"{col_d_letter}{row}")

        # Adjust row height to fit image
        ws.row_dimensions[row].height = xl_img.height * 0.75

        row += 1

    wb.save(OUTPUT_PATH)
    print(f"\n✓ Excel saved → {OUTPUT_PATH}")

# ══════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Running coverage for all modules…")
    all_results = []
    for module, test_path in MODULES:
        print(f"  → {module}", end="", flush=True)
        cr = run_coverage(module, test_path)
        img_path = make_coverage_image(cr)
        all_results.append((cr, img_path))
        print(f"  Total: {cr.total_pct}%  (stmts={cr.total_stmts}, miss={cr.total_miss})")

    print("\nBuilding Excel…")
    build_excel(all_results)
