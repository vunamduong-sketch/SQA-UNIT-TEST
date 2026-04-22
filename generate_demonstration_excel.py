"""
Generate DEMONSTRATION sheet Excel — Domain | Evidence (Terminal Output)
Styled to look like pytest terminal output (dark background, green PASSED text)
"""
import subprocess
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
VENV_PYTHON = "/Users/nguynbon03/Desktop/Combine SQA and Test/.venv/bin/python"
AGODA_BE = "/Users/nguynbon03/Desktop/Combine SQA and Test/SQA/agoda-be"
OUTPUT_PATH = "/Users/nguynbon03/Desktop/Combine SQA and Test/SQA-UNIT-TEST/agoda-be-demonstration.xlsx"

# ── Modules to test ────────────────────────────────────────────────────────
MODULES = [
    ("Accounts",    "accounts/tests.py"),
    ("Activities",  "activities/tests.py"),
    ("Airlines",    "airlines/tests.py"),
    ("Cars",        "cars/tests.py"),
    ("Chatbots",    "chatbots/tests.py"),
    ("Chats",       "chats/tests.py"),
    ("Handbooks",   "handbooks/tests.py"),
    ("Hotels",      "hotels/tests.py"),
    ("Images",      "images/tests.py"),
    ("Payments",    "payments/tests.py"),
    ("Promotions",  "promotions/tests.py"),
    ("Reviews",     "reviews/tests.py"),
    ("Rooms",       "rooms/tests.py"),
]

# ── Colors ─────────────────────────────────────────────────────────────────
CLR_NAVY       = "1F2D3D"   # header background
CLR_AMBER      = "FFC000"   # domain cell
CLR_TERMINAL   = "1E1E2E"   # terminal dark background
CLR_WHITE      = "FFFFFF"
CLR_GREEN      = "4EC94E"   # PASSED green
CLR_CYAN       = "56B6C2"   # separator line cyan

# ── Helpers ────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color="FFFFFF", bold=False, size=10, name="Courier New"):
    return Font(color=hex_color, bold=bold, size=size, name=name)

def border_all(color="AAAAAA"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def run_pytest(test_path: str) -> str:
    """Run pytest for one module, return plain-text output (no ANSI)."""
    env = os.environ.copy()
    env["DJANGO_SETTINGS_MODULE"] = "agoda_be.test_settings"
    env["PYTHONPATH"] = AGODA_BE
    result = subprocess.run(
        [
            VENV_PYTHON, "-m", "pytest",
            test_path,
            "-v", "--no-header", "--tb=short",
            "--color=no",
        ],
        cwd=AGODA_BE,
        capture_output=True,
        text=True,
        env=env,
    )
    return (result.stdout + result.stderr).strip()

# ── Collect outputs ────────────────────────────────────────────────────────
print("Running pytest per module…")
results = []
for domain, path in MODULES:
    print(f"  → {domain}", end="", flush=True)
    output = run_pytest(path)
    results.append((domain, path, output))
    # quick summary
    last = output.splitlines()[-1] if output else ""
    print(f"  {last}")

# ── Build Excel ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "DEMONSTRATION"

# ── Sheet header row ───────────────────────────────────────────────────────
ws.row_dimensions[1].height = 28
for col, label in enumerate(["Domain", "Evidence (pytest Terminal Output)"], 1):
    cell = ws.cell(row=1, column=col, value=label)
    cell.font = Font(color=CLR_WHITE, bold=True, size=12, name="Calibri")
    cell.fill = fill(CLR_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all("000000")

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 110

# ── Data rows (one row per module) ────────────────────────────────────────
for idx, (domain, path, output) in enumerate(results):
    row = idx + 2

    # ── Domain cell ───────────────────────────────────────────────────────
    d_cell = ws.cell(row=row, column=1, value=domain)
    d_cell.font = Font(color="000000", bold=True, size=11, name="Calibri")
    d_cell.fill = fill(CLR_AMBER)
    d_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    d_cell.border = border_all("000000")

    # ── Evidence cell — terminal-styled rich text via plain text + bg ─────
    lines = output.splitlines()

    # Build display text: replace PASSED marker for clarity
    display_lines = []
    for line in lines:
        display_lines.append(line)
    evidence_text = "\n".join(display_lines)

    e_cell = ws.cell(row=row, column=2, value=evidence_text)
    e_cell.font = font(hex_color=CLR_WHITE, size=9)
    e_cell.fill = fill(CLR_TERMINAL)
    e_cell.alignment = Alignment(
        horizontal="left", vertical="top",
        wrap_text=True
    )
    e_cell.border = border_all("333333")

    # Row height: ~14px per line
    line_count = len(display_lines)
    ws.row_dimensions[row].height = max(60, line_count * 13)

# ── Save ──────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\n✓ Excel saved → {OUTPUT_PATH}")
