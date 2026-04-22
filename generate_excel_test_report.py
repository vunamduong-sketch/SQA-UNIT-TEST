"""
Script xuất báo cáo test matrix cho agoda-be ra file Excel.
Output: agoda-be-test-report.xlsx (2 sheets: Scope of Testing + Unit Test Cases)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "agoda-be-test-report.xlsx"

# Màu sắc
HEADER_BG  = "1F2D3D"   # navy dark
HEADER_FG  = "FFFFFF"   # white
GROUP_BG   = "FFC000"   # amber
GROUP_FG   = "000000"   # black
PASS_COLOR = "00B050"   # green

thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_style(cell):
    cell.font = Font(bold=True, color=HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def group_style(cell):
    cell.font = Font(bold=True, color=GROUP_FG)
    cell.fill = PatternFill("solid", fgColor=GROUP_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border


def data_style(cell, wrap=True):
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    cell.border = border


def pass_style(cell):
    cell.font = Font(bold=True, color=PASS_COLOR)
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.border = border


wb = Workbook()

# ============================================================
# SHEET 1: Scope of Testing
# ============================================================
ws1 = wb.active
ws1.title = "Scope of Testing"

headers1 = ["No.", "File / Class", "Path", "Function / Method", "Reason"]
col_widths1 = [6, 28, 45, 28, 55]

ws1.row_dimensions[1].height = 30
for c, (h, w) in enumerate(zip(headers1, col_widths1), 1):
    cell = ws1.cell(row=1, column=c, value=h)
    header_style(cell)
    ws1.column_dimensions[get_column_letter(c)].width = w

scope_rows = [
    (True,  None, "agoda-be\\accounts", "", "", ""),
    (False, 1,  "CustomUser",         "agoda-be\\accounts\\models.py",     "__str__",                    "Returns username; core identity shown across UI/logs."),
    (False, 2,  "RegisterSerializer", "agoda-be\\accounts\\serializers.py","create",                     "Creates user with hashed password; core registration flow."),
    (False, 3,  "CustomUser",         "agoda-be\\accounts\\models.py",     "manager FK",                 "Manager FK can be assigned; drives hotel staff hierarchy."),

    (True,  None, "agoda-be\\activities", "", "", ""),
    (False, 4,  "Activity",           "agoda-be\\activities\\models.py",   "save",                       "Recomputes weighted score on every save; affects ranking/recommendation output."),
    (False, 5,  "UserActivityInteraction","agoda-be\\activities\\models.py","update_weighted_score",    "Updates personalized user-activity score for recommendation logic."),
    (False, 6,  "ActivityDate",       "agoda-be\\activities\\models.py",   "get_active_promotion",       "Resolves active promotion and best discount for pricing."),
    (False, 7,  "ActivityDateBookingDetail","agoda-be\\activities\\models.py","save",                   "Calculates total/final price, updates participants, propagates booking totals."),

    (True,  None, "agoda-be\\airlines", "", "", ""),
    (False, 8,  "Airline",            "agoda-be\\airlines\\models.py",     "__str__",                    "Ensures airline display text is correct (name + code)."),
    (False, 9,  "Aircraft",           "agoda-be\\airlines\\models.py",     "__str__",                    "Ensures aircraft display text includes airline code/model/registration."),
    (False, 10, "Aircraft",           "agoda-be\\airlines\\models.py",     "seat counts",                "Economy/business/first-class seat counts are stored and retrieved correctly."),
    (False, 11, "Aircraft",           "agoda-be\\airlines\\models.py",     "is_active",                  "is_active defaults to True; aircraft available from creation."),

    (True,  None, "agoda-be\\cars", "", "", ""),
    (False, 12, "Car",                "agoda-be\\cars\\models.py",         "calc_total_weighted_score",  "Core ranking formula for car recommendation ordering."),
    (False, 13, "Car",                "agoda-be\\cars\\models.py",         "update_total_weighted_score","Persists weighted score used in sorting/filtering."),
    (False, 14, "Car",                "agoda-be\\cars\\models.py",         "get_active_promotion",       "Finds valid best promotion before price calculation."),
    (False, 15, "CarBookingDetail",   "agoda-be\\cars\\models.py",         "save",                       "Assigns driver, updates driver status, computes total/final price, syncs booking totals."),
    (False, 16, "UserCarInteraction", "agoda-be\\cars\\models.py",         "update_weighted_score",      "Updates user-car affinity score for personalization."),

    (True,  None, "agoda-be\\chatbots", "", "", ""),
    (False, 17, "(No custom model)",  "agoda-be\\chatbots\\models.py",     "N/A",                        "Module currently has no model class/method to test."),

    (True,  None, "agoda-be\\chats", "", "", ""),
    (False, 18, "Conversation",       "agoda-be\\chats\\models.py",        "__str__",                    "Verifies readable conversation label for admin/debugging."),
    (False, 19, "Message",            "agoda-be\\chats\\models.py",        "__str__",                    "Verifies message label includes sender and conversation id."),
    (False, 20, "Message",            "agoda-be\\chats\\models.py",        "seen",                       "seen field defaults to False; ensures unread state on creation."),
    (False, 21, "Conversation",       "agoda-be\\chats\\models.py",        "seen",                       "seen field defaults to False; new conversation starts as unread."),

    (True,  None, "agoda-be\\handbooks", "", "", ""),
    (False, 22, "Handbook",           "agoda-be\\handbooks\\models.py",    "save",                       "Recomputes weighted score; impacts handbook ranking."),
    (False, 23, "Handbook",           "agoda-be\\handbooks\\models.py",    "update_total_weighted_score","Persists aggregate score for recommendation/filtering."),
    (False, 24, "UserHandbookInteraction","agoda-be\\handbooks\\models.py","update_weighted_score",     "Updates user-handbook personalized relevance."),

    (True,  None, "agoda-be\\hotels", "", "", ""),
    (False, 25, "Hotel",              "agoda-be\\hotels\\models.py",       "save",                       "Keeps weighted score consistent for hotel ranking."),
    (False, 26, "Hotel",              "agoda-be\\hotels\\models.py",       "update_min_price",           "Calculates displayed minimum available room price."),
    (False, 27, "UserHotelInteraction","agoda-be\\hotels\\models.py",     "update_weighted_score",      "Updates personalized hotel score from interactions."),

    (True,  None, "agoda-be\\images", "", "", ""),
    (False, 28, "Image",              "agoda-be\\images\\models.py",       "__str__",                    "Ensures stored file name/path is rendered correctly."),
    (False, 29, "Image",             "agoda-be\\images\\models.py",        "uploaded_at",                "uploaded_at is auto-set on creation; tracks upload timestamp."),

    (True,  None, "agoda-be\\payments", "", "", ""),
    (False, 30, "Payment",            "agoda-be\\payments\\models.py",     "status default",             "Critical payment flow default state must start as Pending."),
    (False, 31, "Payment",            "agoda-be\\payments\\models.py",     "method/amount",              "Payment method and amount are stored correctly."),
    (False, 32, "Payment",            "agoda-be\\payments\\models.py",     "transaction_id",             "Optional transaction_id can be stored for payment tracking."),
    (False, 33, "Payment",            "agoda-be\\payments\\models.py",     "multiple payments",          "Multiple payments for same booking are independent records."),

    (True,  None, "agoda-be\\promotions", "", "", ""),
    (False, 34, "Promotion",          "agoda-be\\promotions\\models.py",   "__str__",                    "Verifies promotion label with promotion type display."),
    (False, 35, "RoomPromotion",      "agoda-be\\promotions\\models.py",   "__str__",                    "Ensures room-promotion linkage text is correct."),
    (False, 36, "CarPromotion",       "agoda-be\\promotions\\models.py",   "__str__",                    "Ensures car-promotion linkage text handles nullable car and named car."),
    (False, 37, "ActivityPromotion",  "agoda-be\\promotions\\models.py",   "__str__",                    "Ensures activity date linkage is traceable for debugging."),

    (True,  None, "agoda-be\\reviews", "", "", ""),
    (False, 38, "Review",             "agoda-be\\reviews\\models.py",      "service_type_name",          "Returns service label used in response and display logic."),
    (False, 39, "Review",             "agoda-be\\reviews\\models.py",      "get_service_instance",       "Maps review to concrete service object (hotel/activity)."),
    (False, 40, "Review",             "agoda-be\\reviews\\models.py",      "__str__",                    "Produces readable review identity containing service/ref id."),

    (True,  None, "agoda-be\\rooms", "", "", ""),
    (False, 41, "Room",               "agoda-be\\rooms\\models.py",        "capacity",                   "Computes total room capacity from adults + children."),
    (False, 42, "Room",               "agoda-be\\rooms\\models.py",        "save",                       "Updates availability state and triggers hotel min price refresh."),
    (False, 43, "Room",               "agoda-be\\rooms\\models.py",        "decrease_available_rooms",   "Prevents negative inventory and updates availability after booking."),
    (False, 44, "Room",               "agoda-be\\rooms\\models.py",        "get_active_promotion",       "Finds current best room-level promotion for final pricing."),
    (False, 45, "RoomBookingDetail",  "agoda-be\\rooms\\models.py",        "save",                       "Calculates booking amount (overnight/dayuse), applies discount, decrements stock, syncs booking totals."),
]

row_num = 2
for entry in scope_rows:
    is_group = entry[0]
    if is_group:
        group_name = entry[2]
        ws1.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        cell = ws1.cell(row=row_num, column=1, value=group_name)
        group_style(cell)
        ws1.row_dimensions[row_num].height = 20
    else:
        _, no, file_class, path, method, reason = entry
        vals = [no, file_class, path, method, reason]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row_num, column=c, value=v)
            data_style(cell)
        ws1.row_dimensions[row_num].height = 25
    row_num += 1

# ============================================================
# SHEET 2: Unit Test Cases
# ============================================================
ws2 = wb.create_sheet("Unit Test Cases")

headers2 = ["TC ID", "File / Class", "Path", "Function / Method", "Test Objective", "Input", "Expected Output", "Result", "Notes"]
col_widths2 = [14, 26, 42, 26, 38, 42, 42, 9, 20]

ws2.row_dimensions[1].height = 30
for c, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=c, value=h)
    header_style(cell)
    ws2.column_dimensions[get_column_letter(c)].width = w

tc_rows = [
    # ACCOUNTS
    (True,  "agoda-be\\accounts", "", "", "", "", "", "", "", ""),
    (False, "ACC-TC-001", "CustomUser",         "agoda-be\\accounts\\models.py",     "__str__",
        "Returns username as string representation",
        "Create CustomUser with username='acc_user'",
        "str(user) == 'acc_user'", "Pass", ""),
    (False, "ACC-TC-002", "CustomUser",         "agoda-be\\accounts\\models.py",     "defaults",
        "Role defaults to 'customer', driver_status defaults to 'idle'",
        "Create CustomUser without specifying role or driver_status",
        "user.role == 'customer'\nuser.driver_status == 'idle'", "Pass", ""),
    (False, "ACC-TC-003", "CustomUser",         "agoda-be\\accounts\\models.py",     "assign driver",
        "Role and driver_status can be assigned 'driver' / 'busy'",
        "Create CustomUser with role='driver', driver_status='busy'",
        "user.role == 'driver'\nuser.driver_status == 'busy'", "Pass", ""),
    (False, "ACC-TC-004", "RegisterSerializer", "agoda-be\\accounts\\serializers.py","create",
        "Creates user with correct fields and hashed password",
        "Call RegisterSerializer with valid data dict (username, email, password, ...)",
        "- User saved in DB\n- password is hashed (not plaintext)\n- check_password() returns True", "Pass", ""),
    (False, "ACC-TC-005", "CustomUser",         "agoda-be\\accounts\\models.py",     "manager FK",
        "manager FK can be set to another CustomUser",
        "Create manager user, assign as manager when creating staff user",
        "staff.manager == manager", "Pass", ""),

    # ACTIVITIES
    (True,  "agoda-be\\activities", "", "", "", "", "", "", "", ""),
    (False, "ACT-TC-001", "Activity",                "agoda-be\\activities\\models.py", "save",
        "save recomputes total_weighted_score",
        "Create Activity with avg_star=4.5, total_click=9, total_positive=8, total_negative=2",
        "activity.total_weighted_score == activity.calc_total_weighted_score", "Pass", ""),
    (False, "ACT-TC-002", "UserActivityInteraction", "agoda-be\\activities\\models.py", "update_weighted_score",
        "Updates personalized user-activity score > 0",
        "Create interaction: click_count=3, positive_count=4, negative_count=1, neutral_count=1\nCall update_weighted_score()",
        "interaction.weighted_score > 0", "Pass", ""),
    (False, "ACT-TC-003", "ActivityDate",             "agoda-be\\activities\\models.py", "get_active_promotion",
        "Returns best active discount (discount_percent=15)",
        "Create ActivityPromotion: is_active=True, within date range, discount_percent=15",
        "get_active_promotion()['discount_percent'] == 15", "Pass", ""),
    (False, "ACT-TC-004", "ActivityDateBookingDetail","agoda-be\\activities\\models.py","save",
        "Calculates total/final price, decrements participants, syncs booking total",
        "price_adult=300, adult_qty=2, price_child=120, child_qty=1",
        "- total_price == 720.0\n- final_price == 720.0\n- participants_available decremented by 3\n- booking.total_price == 720.0", "Pass", ""),

    # AIRLINES
    (True,  "agoda-be\\airlines", "", "", "", "", "", "", "", ""),
    (False, "AIRL-TC-001", "Airline",  "agoda-be\\airlines\\models.py", "__str__",
        "Returns name and code in format 'Name (CODE)'",
        "Create Airline(name='Vietnam Airlines', code='VN')",
        "str(airline) == 'Vietnam Airlines (VN)'", "Pass", ""),
    (False, "AIRL-TC-002", "Aircraft", "agoda-be\\airlines\\models.py", "__str__",
        "Returns airline code, model, registration number",
        "Create Airline(code='VJ'), Aircraft(model='Airbus A321', registration_number='VN-A123')",
        "str(aircraft) == 'VJ - Airbus A321 (VN-A123)'", "Pass", ""),
    (False, "AIRL-TC-003", "Aircraft", "agoda-be\\airlines\\models.py", "seat counts",
        "Stores economy/business/first_class seat counts correctly",
        "Create Aircraft: economy_seats=200, business_seats=28, first_class_seats=8, total_seats=236",
        "All seat count fields retrieved match input values", "Pass", ""),
    (False, "AIRL-TC-004", "Aircraft", "agoda-be\\airlines\\models.py", "is_active",
        "is_active defaults to True on creation",
        "Create Aircraft without specifying is_active",
        "aircraft.is_active == True", "Pass", ""),

    # CARS
    (True,  "agoda-be\\cars", "", "", "", "", "", "", "", ""),
    (False, "CAR-TC-001", "Car",               "agoda-be\\cars\\models.py", "calc_total_weighted_score",
        "calc_total_weighted_score == total_booking_count",
        "Create Car with total_booking_count=12",
        "car.calc_total_weighted_score == 12", "Pass", ""),
    (False, "CAR-TC-002", "Car",               "agoda-be\\cars\\models.py", "update_total_weighted_score",
        "Persists weighted score to DB",
        "Create Car, call update_total_weighted_score(), refresh_from_db()",
        "car.total_weighted_score == 12", "Pass", ""),
    (False, "CAR-TC-003", "Car",               "agoda-be\\cars\\models.py", "get_active_promotion",
        "Returns highest active CarPromotion discount",
        "Create CarPromotion: is_active=True, within date range, discount_percent=20",
        "get_active_promotion()['discount_percent'] == 20", "Pass", ""),
    (False, "CAR-TC-004", "CarBookingDetail",  "agoda-be\\cars\\models.py", "save",
        "Assigns driver, computes total price, sets driver status to busy",
        "Create CarBookingDetail: car.price_per_km=10000, distance_km=10",
        "- driver assigned\n- total_price == 200000.0\n- driver.driver_status == 'busy'\n- booking.total_price synced", "Pass", ""),
    (False, "CAR-TC-005", "CarBookingDetail",  "agoda-be\\cars\\models.py", "save (ARRIVED)",
        "Sets driver status back to idle when booking status is ARRIVED",
        "Create CarBookingDetail with status=CarBookingStatus.ARRIVED",
        "driver.driver_status == 'idle'", "Pass", ""),
    (False, "CAR-TC-006", "UserCarInteraction","agoda-be\\cars\\models.py", "update_weighted_score",
        "weighted_score == booking_count",
        "Create UserCarInteraction: booking_count=4, call update_weighted_score()",
        "interaction.weighted_score == 4", "Pass", ""),

    # CHATBOTS
    (True,  "agoda-be\\chatbots", "", "", "", "", "", "", "", ""),
    (False, "CHATB-TC-001", "(No model)", "agoda-be\\chatbots\\models.py", "N/A",
        "chatbots module has no custom Model class",
        "Inspect chatbot_models for Model subclasses using vars()",
        "model_classes == []", "Pass", ""),

    # CHATS
    (True,  "agoda-be\\chats", "", "", "", "", "", "", "", ""),
    (False, "CHAT-TC-001", "Conversation", "agoda-be\\chats\\models.py", "__str__",
        "Readable label with both user names",
        "Create Conversation(user1='chat_user_1', user2='chat_user_2')",
        "'Conversation between chat_user_1 and chat_user_2' in str(conv)", "Pass", ""),
    (False, "CHAT-TC-002", "Message", "agoda-be\\chats\\models.py", "__str__",
        "Label includes sender username and conversation UUID",
        "Create Message from chat_user_1 in conversation",
        "'Message from chat_user_1' and str(conv.id) in str(message)", "Pass", ""),
    (False, "CHAT-TC-003", "Message", "agoda-be\\chats\\models.py", "seen",
        "seen field defaults to False on creation",
        "Create Message without specifying seen",
        "message.seen == False", "Pass", ""),
    (False, "CHAT-TC-004", "Conversation", "agoda-be\\chats\\models.py", "seen",
        "Conversation.seen defaults to False on creation",
        "Create Conversation without specifying seen",
        "conversation.seen == False", "Pass", ""),

    # HANDBOOKS
    (True,  "agoda-be\\handbooks", "", "", "", "", "", "", "", ""),
    (False, "HB-TC-001", "Handbook",               "agoda-be\\handbooks\\models.py", "save",
        "save recomputes total_weighted_score",
        "Create Handbook with total_click=20, avg_star=4.0, total_positive=10",
        "handbook.total_weighted_score == handbook.calc_total_weighted_score", "Pass", ""),
    (False, "HB-TC-002", "Handbook",               "agoda-be\\handbooks\\models.py", "update_total_weighted_score",
        "Persists aggregate score to DB",
        "Create Handbook, set total_click=5, call update_total_weighted_score(), refresh",
        "handbook.total_weighted_score == handbook.calc_total_weighted_score", "Pass", ""),
    (False, "HB-TC-003", "UserHandbookInteraction", "agoda-be\\handbooks\\models.py","update_weighted_score",
        "Personalized relevance score > 0",
        "Create interaction: click_count=4, positive_count=3, negative_count=1\nCall update_weighted_score()",
        "interaction.weighted_score > 0", "Pass", ""),

    # HOTELS
    (True,  "agoda-be\\hotels", "", "", "", "", "", "", "", ""),
    (False, "HTL-TC-001", "Hotel",                "agoda-be\\hotels\\models.py", "save",
        "save recomputes total_weighted_score",
        "Create Hotel: avg_star=4.2, total_click=10, total_positive=9, total_negative=1",
        "hotel.total_weighted_score == hotel.calc_total_weighted_score", "Pass", ""),
    (False, "HTL-TC-002", "Hotel",                "agoda-be\\hotels\\models.py", "update_min_price",
        "Uses only rooms where available=True",
        "Create 2 rooms: one price=100/available, one price=300/unavailable\nCall update_min_price()",
        "hotel.min_price == 100", "Pass", ""),
    (False, "HTL-TC-003", "UserHotelInteraction", "agoda-be\\hotels\\models.py", "update_weighted_score",
        "Personalized hotel score > 0",
        "Create interaction: click_count=5, positive_count=4, negative_count=1\nCall update_weighted_score()",
        "interaction.weighted_score > 0", "Pass", ""),

    # IMAGES
    (True,  "agoda-be\\images", "", "", "", "", "", "", "", ""),
    (False, "IMG-TC-001", "Image", "agoda-be\\images\\models.py", "__str__",
        "Returns image file name/path",
        "Create Image(image='uploads/images/test-image.jpg')",
        "str(image) == 'uploads/images/test-image.jpg'", "Pass", ""),
    (False, "IMG-TC-002", "Image", "agoda-be\\images\\models.py", "uploaded_at",
        "uploaded_at is auto-set on creation",
        "Create Image, record time before/after",
        "image.uploaded_at is not None\nbefore <= uploaded_at <= after", "Pass", ""),

    # PAYMENTS
    (True,  "agoda-be\\payments", "", "", "", "", "", "", "", ""),
    (False, "PAY-TC-001", "Payment", "agoda-be\\payments\\models.py", "status default",
        "Payment defaults to PENDING status",
        "Create Payment without specifying status",
        "payment.status == PaymentStatus.PENDING", "Pass", ""),
    (False, "PAY-TC-002", "Payment", "agoda-be\\payments\\models.py", "method/amount",
        "Stores payment method and amount correctly",
        "Create Payment with method=CASH, amount=320000, status=SUCCESS",
        "payment.method == CASH\npayment.amount == 320000\npayment.status == SUCCESS", "Pass", ""),
    (False, "PAY-TC-003", "Payment", "agoda-be\\payments\\models.py", "transaction_id",
        "Can store optional transaction_id",
        "Create Payment with transaction_id='TXN-AGODA-123456'",
        "payment.transaction_id == 'TXN-AGODA-123456'", "Pass", ""),
    (False, "PAY-TC-004", "Payment", "agoda-be\\payments\\models.py", "multiple payments",
        "Multiple payments for same booking are independent",
        "Create 2 Payments for same booking with different amounts/status",
        "- payment1.amount == 500000, status == PENDING\n- payment2.amount == 250000, status == SUCCESS", "Pass", ""),

    # PROMOTIONS
    (True,  "agoda-be\\promotions", "", "", "", "", "", "", "", ""),
    (False, "PRM-TC-001", "Promotion",         "agoda-be\\promotions\\models.py", "__str__",
        "Returns title and promotion type display name",
        "Create Promotion(title='Mega Sale', type=HOTEL)",
        "str(promo) == 'Mega Sale (Chỗ ở)'", "Pass", ""),
    (False, "PRM-TC-002", "RoomPromotion",     "agoda-be\\promotions\\models.py", "__str__",
        "Shows promotion title and room type",
        "Create RoomPromotion with room.room_type='Suite'",
        "'Suite' in str(room_promo)\n'Mega Sale' in str(room_promo)", "Pass", ""),
    (False, "PRM-TC-003", "CarPromotion",      "agoda-be\\promotions\\models.py", "__str__",
        "Handles nullable car → shows 'N/A'",
        "Create CarPromotion(car=None)",
        "'N/A' in str(car_promo)", "Pass", ""),
    (False, "PRM-TC-004", "ActivityPromotion", "agoda-be\\promotions\\models.py", "__str__",
        "Contains activity date id for traceability",
        "Create ActivityPromotion with activity_date",
        "f'ActivityDate: {activity_date.id}' in str(activity_promo)", "Pass", ""),
    (False, "PRM-TC-005", "CarPromotion",      "agoda-be\\promotions\\models.py", "__str__ (with car)",
        "Shows car name when car is not null",
        "Create CarPromotion with car.name='Promo Car'",
        "'Promo Car' in str(car_promo)", "Pass", ""),

    # REVIEWS
    (True,  "agoda-be\\reviews", "", "", "", "", "", "", "", ""),
    (False, "RVW-TC-001", "Review", "agoda-be\\reviews\\models.py", "service_type_name",
        "Returns service label (e.g. 'Hotel')",
        "Create Review(service_type=ServiceType.HOTEL)",
        "review.service_type_name == 'Hotel'", "Pass", ""),
    (False, "RVW-TC-002", "Review", "agoda-be\\reviews\\models.py", "get_service_instance",
        "Maps review to Hotel or Activity object",
        "- Create hotel_review (service_type=HOTEL)\n- Create activity_review (service_type=ACTIVITY)",
        "- hotel_review.get_service_instance() == hotel\n- activity_review.get_service_instance() == activity", "Pass", ""),
    (False, "RVW-TC-003", "Review", "agoda-be\\reviews\\models.py", "__str__",
        "Contains service type label and ref id",
        "Create Review(service_type=HOTEL, service_ref_id=hotel.id)",
        "'Hotel' in str(review)\nf'#{hotel.id}' in str(review)", "Pass", ""),

    # ROOMS
    (True,  "agoda-be\\rooms", "", "", "", "", "", "", "", ""),
    (False, "ROOM-TC-001", "Room",              "agoda-be\\rooms\\models.py", "capacity",
        "Returns adults_capacity + children_capacity",
        "Create Room(adults_capacity=2, children_capacity=1)",
        "room.capacity == 3", "Pass", ""),
    (False, "ROOM-TC-002", "Room",              "agoda-be\\rooms\\models.py", "save",
        "Marks room unavailable when available_rooms=0",
        "Create Room, set available_rooms=0, call save()",
        "room.available == False", "Pass", ""),
    (False, "ROOM-TC-003", "Room",              "agoda-be\\rooms\\models.py", "decrease_available_rooms",
        "Cannot decrement below 0",
        "Room has available_rooms=3, call decrease_available_rooms(5)",
        "room.available_rooms == 0", "Pass", ""),
    (False, "ROOM-TC-004", "Room",              "agoda-be\\rooms\\models.py", "get_active_promotion",
        "Returns best active RoomPromotion",
        "Create RoomPromotion: is_active=True, within date range, discount_percent=25",
        "get_active_promotion()['discount_percent'] == 25", "Pass", ""),
    (False, "ROOM-TC-005", "RoomBookingDetail", "agoda-be\\rooms\\models.py", "save (overnight)",
        "Calculates overnight price: nights x rooms x price_per_night",
        "price_per_night=100, 2 nights, 2 rooms",
        "- detail.total_price == 400.0\n- detail.final_price == 400.0\n- available_rooms decremented\n- booking.total_price == 400.0", "Pass", ""),
    (False, "ROOM-TC-006", "RoomBookingDetail", "agoda-be\\rooms\\models.py", "save (dayuse)",
        "Calculates dayuse price: price_per_day x room_count",
        "dayuse room: price_per_day=150, room_count=1, check_out 6h after check_in",
        "detail.total_price == 150.0", "Pass", ""),
]

row_num = 2
for entry in tc_rows:
    is_group = entry[0]
    if is_group:
        group_name = entry[1]
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        cell = ws2.cell(row=row_num, column=1, value=group_name)
        group_style(cell)
        ws2.row_dimensions[row_num].height = 20
    else:
        _, tc_id, file_class, path, method, objective, input_desc, expected, result, notes = entry
        vals = [tc_id, file_class, path, method, objective, input_desc, expected, result, notes]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_num, column=c, value=v)
            if c == 8:  # Result column
                pass_style(cell)
            else:
                data_style(cell)
        ws2.row_dimensions[row_num].height = 40
    row_num += 1

wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE}")
